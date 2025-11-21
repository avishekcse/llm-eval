import os
import openai
from openpyxl import Workbook, load_workbook
import glob

# ===== STEP 0: Set your API key here =====
openai.api_key = "your_api_key"  # Replace with your OpenAI API key

# ===== STEP 1: Upload multiple files =====
def upload_files_from_folder(folder_path):
    """Uploads all text files from the given folder to OpenAI."""
    file_ids = []
    for file_path in glob.glob(os.path.join(folder_path, "*.txt")):
        with open(file_path, "rb") as f:
            response = openai.File.create(file=f, purpose="answers")  # Use 'answers' for querying
            print(f"Uploaded: {file_path} | File ID: {response['id']}")
            file_ids.append(response['id'])
    return file_ids


# ===== STEP 2: Ask one question about all files =====
def ask_one_question(file_ids, question, model="gpt-4"):
    """Ask a single question about all uploaded files."""
    print(f"\n🔹 Asking: {question}")
    
    response = openai.Completion.create(
        model=model,
        prompt=question,
        files=file_ids  # Send file IDs for analysis
    )
    
    answer = response.choices[0].text.strip()
    print("Answer received.")
    return answer


# ===== STEP 3: Read Failure Modes from Excel =====
def read_failure_modes_from_excel(excel_path):
    """Reads failure mode data from an Excel file."""
    failure_modes = []
    wb = load_workbook(excel_path)
    ws = wb.active

    # Assuming the failure modes are in the first sheet and the columns are: 
    # Failure Mode, Example, Document Structure & Context, Chunking & Pre-processing Strategy, Embedding & Retrieval Strategy
    for row in ws.iter_rows(min_row=2, values_only=True):  # Start from row 2 to skip header
        failure_mode = {
            "Failure Mode": row[0],
            "Example": row[1],
            "Document Structure & Context": row[2],
            "Chunking & Pre-processing Strategy": row[3],
            "Embedding & Retrieval Strategy": row[4]
        }
        failure_modes.append(failure_mode)
    
    return failure_modes


# ===== STEP 4: Save result to Excel =====
def save_to_excel(question, answer, filename="generated_results.xlsx"):
    """Appends Q&A to an existing Excel file."""
    try:
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        # If file doesn't exist, create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "ChatGPT Q&A"
        # Write headers only if file is new
        ws.append(["Question", "Answer"])
    
    # Append the question and answer
    ws.append([question, answer])

    # Format columns
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 120

    wb.save(filename)
    print(f"Saved result to Excel: {filename}")


# ===== MAIN EXECUTION =====
if __name__ == "__main__":
    folder_path = "./source_folder"  # Folder with your text files
    excel_path = "./input_sample_failure_mode.xlsx"  # Path to the failure mode Excel file

    # Read failure modes from the Excel file
    failure_modes = read_failure_modes_from_excel(excel_path)
    print(f"Read {len(failure_modes)} failure modes from {excel_path}.")

    # Upload all files from the source folder
    file_ids = upload_files_from_folder(folder_path)

    # Prepare list of questions
    questions = [
        "Summarize the content of all uploaded text files.",
        "Extract key entity types from all files and cluster them using K-means. Use silhouette coefficient to determine the optimal number of clusters.",
        "Extract relations between entities in the uploaded text files. Form triples (Subject, Predicate, Object) and aggregate them into a Knowledge Graph.",
        "Store the extracted triples in a CSV file.",
        "Generate similar failure mode examples based on the uploaded text files. Output should include: Failure Mode, Example, Document Structure & Context, Chunking & Pre-processing Strategy, Embedding & Retrieval Strategy."
    ]

    # Ask each question and save result
    for question in questions:
        answer = ask_one_question(file_ids, question)
        save_to_excel(question, answer, "generated_results.xlsx")

    print("\nProcess complete! Results saved in generated_results.xlsx")

